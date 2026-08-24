#!/usr/bin/env python3
# ============================================================
# Gate — extrai os vetores da planilha de pontuação paralela do Kimi
# (camada 2) para JSON, que o runner tsx (scripts/gate-run.ts) roda
# pelas funções REAIS do app e cruza célula a célula.
#
# Entrada:  gate/planilha_pontuacao_paralela_gate.xlsx
# Saída:    gate/vetores-b.json  { vetoresB: [...], v17: [...] }
#
# Rodar (precisa openpyxl):  python3 scripts/extract-gate-vectors.py
# (O JSON já vai versionado; só re-extrair se a planilha mudar.)
# ============================================================
import json, os, sys, re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "gate", "planilha_pontuacao_paralela_gate.xlsx")
OUT = os.path.join(ROOT, "gate", "vetores-b.json")

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ----- Vetores: escore/comportamento esperados + banco -----
V = wb["Vetores"]
meta = {}
for r in list(V.iter_rows(values_only=True))[1:]:
    vid = r[0]
    if not vid:
        continue
    banco = str(r[2]).strip().lower() if r[2] is not None else ""
    meta[vid] = {
        "usesDatabase": banco == "sim",
        "escoreEsperado": r[4],          # int, "—", ou combinado (V16)
        "comportamento": (str(r[5]).strip() if r[5] is not None else ""),
    }

# ----- Entradas_B: respostas por vetor -----
E = wb["Entradas_B"]
rows = list(E.iter_rows(values_only=True))
hdr = rows[0]
vec_cols = list(hdr[2:])

def norm_matriz(v):
    if v is None:
        return None
    s = str(v).strip()
    m = {"Sim": "sim", "Não": "nao", "N/A": "na"}
    if s in m:
        return m[s]
    if s.startswith("Parcial"):   # checklist incompleto (P6.b.6) = resposta de risco "não"
        return "nao"
    if s == "(neutro)":
        return None
    return None  # qualquer outro marcador em item de matriz → não respondido

# remapeia id de contexto da planilha (C.1/C.2) para o id da spec (contexto1/contexto2)
CTX_REMAP = {"C.1": "contexto1", "C.2": "contexto2"}

vetoresB = {vid: {"vetor": vid, "usesDatabase": meta.get(vid, {}).get("usesDatabase", False),
                  "quantAnswers": {}, "contextAnswers": {},
                  "escoreEsperado": meta.get(vid, {}).get("escoreEsperado"),
                  "comportamento": meta.get(vid, {}).get("comportamento", "")}
            for vid in vec_cols}

for r in rows[1:]:
    iid = r[0]
    if not iid:
        continue
    iid = str(iid)
    for j, cell in enumerate(r[2:]):
        vid = vec_cols[j]
        if iid.startswith("C."):
            if cell is None or str(cell).strip() in ("(neutro)", ""):
                continue
            cid = CTX_REMAP.get(iid, iid)
            vetoresB[vid]["contextAnswers"][cid] = str(cell).strip()
        else:
            a = norm_matriz(cell)
            if a is not None:
                vetoresB[vid]["quantAnswers"][iid] = a

# ----- V17 (Versão A): contagem de risco por eixo → nível esperado -----
EIXO_ID = {"1": "eixo1", "2": "eixo2", "3": "eixo3", "3.b": "eixo3b", "4": "eixo4", "5": "eixo5"}
def eixo_id_from_name(name):
    m = re.search(r"Eixo\s+(3\.b|\d)", str(name))
    return EIXO_ID.get(m.group(1)) if m else None

v17 = []
for r in list(wb["V17_VersaoA"].iter_rows(values_only=True))[1:]:
    linha, eixo, nperg, nrisco, esperado = r[0], r[1], r[2], r[3], r[4]
    if linha is None:
        continue
    eid = eixo_id_from_name(eixo)
    raw = str(nrisco).replace("+", "").strip() if nrisco is not None else ""
    if not eid or not raw.isdigit():
        continue  # linhas de separação/nota (ex.: "—")
    lvl = re.sub(r"[^IV]", "", str(esperado))  # "Nível III" -> "III"
    if lvl not in ("I", "II", "III", "IV"):
        continue
    v17.append({"linha": linha, "eixoId": eid, "riskCount": int(raw), "expectedLevel": lvl})

out = {"vetoresB": [vetoresB[v] for v in vec_cols], "v17": v17}
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"OK: {len(out['vetoresB'])} vetores B + {len(out['v17'])} linhas V17 → {os.path.relpath(OUT, ROOT)}")
